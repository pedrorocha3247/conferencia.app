# -*- coding: utf-8 -*-
import os
import re
import unicodedata
import io
import fitz  # PyMuPDF
import pandas as pd
from collections import OrderedDict
from flask import Flask, request, send_file, url_for, make_response, jsonify, session, redirect
import json
import traceback
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import zipfile
import time # Importado para logs
import base64
import requests
from datetime import datetime

# ==== Constantes e Mapeamentos ====
DASHES = dict.fromkeys(map(ord, "\u2010\u2011\u2012\u2013\u2014\u2015\u2212"), "-")
HEADERS = (
    "Remessa para Conferência", "Página", "Banco", "IMOBILIARIOS", "Débitos do Mês",
    "Vencimento", "Lançamentos", "Programação", "Carta", "DÉBITOS", "ENCARGOS",
    "PAGAMENTO", "TOTAL", "Limite p/", "TOTAL A PAGAR", "PAGAMENTO EFETUADO", "DESCONTO"
)
PADRAO_LOTE = re.compile(r"\b(\d{2,4}\.([A-Z0-9\u0399\u039A]{2})\.\d{1,4})\b")
PADRAO_PARCELA_MESMA_LINHA = re.compile(
    r"^(?!(?:DÉBITOS|ENCARGOS|DESCONTO|PAGAMENTO|TOTAL(?!\s*A PAGAR)|Limite p/))\s*" # <-- MUDANÇA AQUI
    r"([A-Za-zÀ-ú][A-Za-zÀ-ú\s\.\-\/\d]+?)\s+([\d.,]+)"
    r"(?=\s{2,}|\t|$)", re.MULTILINE
)
PADRAO_NUMERO_PURO = re.compile(r"^\s*([\d\.,]+)\s*$")
CODIGO_EMP_MAP = {
    '04': 'RSCI', '05': 'RSCIV', '06': 'RSCII', '07': 'RSCV', '08': 'RSCIII',
    '09': 'IATE', '10': 'MARINA', '11': 'NVI', '12': 'NVII',
    '13': 'SBRRI', '14': 'SBRRII', '15': 'SBRRIII'
}
EMP_MAP = {
    "NVI": {"Melhoramentos": 205.61, "Fundo de Transporte": 9.00},
    "NVII": {"Melhoramentos": 245.47, "Fundo de Transporte": 9.00},
    "RSCI": {"Melhoramentos": 250.42, "Fundo de Transporte": 9.00},
    "RSCII": {"Melhoramentos": 240.29, "Fundo de Transporte": 9.00},
    "RSCIII": {"Melhoramentos": 281.44, "Fundo de Transporte": 9.00},
    "RSCIV": {"Melhoramentos": 324.20, "Fundo de Transporte": 9.00},
    "IATE": {"Melhoramentos": 240.00, "Fundo de Transporte": 9.00},
    "MARINA": {"Melhoramentos": 240.00, "Fundo de Transporte": 9.00},
    "SBRRI": {"Melhoramentos": 245.47, "Fundo de Transporte": 13.00},
    "SBRRII": {"Melhoramentos": 245.47, "Fundo de Transporte": 13.00},
    "SBRRIII": {"Melhoramentos": 245.47, "Fundo de Transporte": 13.00},
    "RSCV": {"Melhoramentos": 280.00, "Fundo de Transporte": 9.00},
}
BASE_FIXOS = {
    "Taxa de Conservação": [434.11],
    "Contrib. Social SLIM": [107.00, 321.00],
    "Contribuição ABRASMA - Bronze": [20.00],
    "Contribuição ABRASMA - Prata": [40.00],
    "Contribuição ABRASMA - Ouro": [60.00],
}
BASE_FIXOS_CCB = {
    "Alienação Fiduciária CCB": [],
    "Financiamento Realiza CCB": [],
    "Encargos Não Pagos CCB": [],
    "Débito por pagamento a menor CCB": [],
    "Crédito por pagamento a maior CCB": [],
    "Negociação Alienação CCB": []
}

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'kasil-validador-chave-interna')
CONFIG_SENHA   = os.environ.get('CONFIG_SENHA', 'kasil2025')

# ==== Integração GitHub (commit automático do config.json) ====
# Configure estas variáveis de ambiente no Render para ativar:
#   GITHUB_TOKEN        -> token de acesso (fine-grained PAT, Contents: Read & Write)
#   GITHUB_REPO         -> "usuario/repositorio" (ex: "pedrorocha3247/conferencia.app")
#   GITHUB_BRANCH       -> branch alvo (padrão: "main")
#   GITHUB_CONFIG_PATH  -> caminho do config.json no repo (padrão: "config.json")
GITHUB_TOKEN       = os.environ.get('GITHUB_TOKEN', '')
GITHUB_REPO        = os.environ.get('GITHUB_REPO', '')
GITHUB_BRANCH      = os.environ.get('GITHUB_BRANCH', 'main')
GITHUB_CONFIG_PATH = os.environ.get('GITHUB_CONFIG_PATH', 'config.json')

def github_configurado() -> bool:
    return bool(GITHUB_TOKEN and GITHUB_REPO)

def commitar_config_github(config: dict, alteracoes: list = None):
    """Comita o config.json no GitHub. Retorna (ok: bool, mensagem: str)."""
    if not github_configurado():
        return False, 'GitHub não configurado (variáveis de ambiente ausentes).'
    api_url = f'https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_CONFIG_PATH}'
    headers = {
        'Authorization': f'Bearer {GITHUB_TOKEN}',
        'Accept': 'application/vnd.github+json',
        'X-GitHub-Api-Version': '2022-11-28',
    }
    try:
        # 1. Busca o SHA atual do arquivo (necessário para atualizar)
        sha = None
        r_get = requests.get(api_url, headers=headers, params={'ref': GITHUB_BRANCH}, timeout=15)
        if r_get.status_code == 200:
            sha = r_get.json().get('sha')
        elif r_get.status_code != 404:
            return False, f'Falha ao consultar GitHub ({r_get.status_code}): {r_get.text[:200]}'

        # 2. Monta a mensagem do commit
        if alteracoes:
            resumo = '; '.join(a['campo'] for a in alteracoes[:3])
            if len(alteracoes) > 3:
                resumo += f' (+{len(alteracoes) - 3})'
            msg = f'Config via painel: {resumo}'
        else:
            msg = 'Atualização de configuração via painel'

        # 3. Envia o novo conteúdo
        conteudo_b64 = base64.b64encode(
            json.dumps(config, ensure_ascii=False, indent=2).encode('utf-8')
        ).decode('utf-8')
        payload = {'message': msg, 'content': conteudo_b64, 'branch': GITHUB_BRANCH}
        if sha:
            payload['sha'] = sha
        r_put = requests.put(api_url, headers=headers, json=payload, timeout=15)
        if r_put.status_code in (200, 201):
            print(f"[GITHUB] config.json comitado em {GITHUB_REPO}@{GITHUB_BRANCH}.")
            return True, 'Configuração versionada no GitHub.'
        return False, f'Falha ao comitar ({r_put.status_code}): {r_put.text[:200]}'
    except Exception as e:
        print(f"[GITHUB] ERRO ao comitar config: {e}")
        return False, str(e)

# Define UPLOAD_FOLDER como um caminho absoluto relativo à raiz do app
UPLOAD_FOLDER_PATH = os.path.join(app.root_path, 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER_PATH
# Cria o diretório usando o caminho absoluto
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
print(f"Pasta de Upload configurada em: {app.config['UPLOAD_FOLDER']}")

CONFIG_PATH = os.path.join(app.root_path, 'config.json')
print(f"[CONFIG] Caminho do config.json: {CONFIG_PATH}")

CONFIG_PADRAO = {
    "EMP_MAP": {
        "NVI":    {"Melhoramentos": 205.61, "Fundo de Transporte": 9.00},
        "NVII":   {"Melhoramentos": 245.47, "Fundo de Transporte": 9.00},
        "RSCI":   {"Melhoramentos": 250.42, "Fundo de Transporte": 9.00},
        "RSCII":  {"Melhoramentos": 240.29, "Fundo de Transporte": 9.00},
        "RSCIII": {"Melhoramentos": 281.44, "Fundo de Transporte": 9.00},
        "RSCIV":  {"Melhoramentos": 324.20, "Fundo de Transporte": 9.00},
        "RSCV":   {"Melhoramentos": 280.00, "Fundo de Transporte": 9.00},
        "IATE":   {"Melhoramentos": 240.00, "Fundo de Transporte": 9.00},
        "MARINA": {"Melhoramentos": 240.00, "Fundo de Transporte": 9.00},
        "SBRRI":  {"Melhoramentos": 245.47, "Fundo de Transporte": 13.00},
        "SBRRII": {"Melhoramentos": 245.47, "Fundo de Transporte": 13.00},
        "SBRRIII":{"Melhoramentos": 245.47, "Fundo de Transporte": 13.00},
    },
    "BASE_FIXOS": {
        "Taxa de Conservação":           [434.11],
        "Contrib. Social SLIM":          [321.00, 107.00],
        "Contribuição ABRASMA - Bronze": [20.00],
        "Contribuição ABRASMA - Prata":  [40.00],
        "Contribuição ABRASMA - Ouro":   [60.00],
    }
}

def carregar_config() -> dict:
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
            return cfg
    except FileNotFoundError:
        print(f"[CONFIG] config.json não encontrado em '{CONFIG_PATH}'. Usando padrão.")
        return {k: dict(v) for k, v in CONFIG_PADRAO.items()}
    except json.JSONDecodeError as e:
        print(f"[CONFIG] ERRO ao ler config.json (JSON inválido): {e}. Usando padrão.")
        return {k: dict(v) for k, v in CONFIG_PADRAO.items()}

HISTORY_PATH = os.path.join(app.root_path, 'config_historico.json')

def carregar_historico() -> list:
    try:
        with open(HISTORY_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

def _fmt_val(v):
    if isinstance(v, list):
        return ' / '.join(f'R$ {float(x):.2f}' for x in v)
    return f'R$ {float(v):.2f}'

def _detectar_alteracoes(antigo: dict, novo: dict) -> list:
    diffs = []
    for emp, vals in novo.get('EMP_MAP', {}).items():
        for campo, novo_val in vals.items():
            ant = antigo.get('EMP_MAP', {}).get(emp, {}).get(campo)
            if ant is not None and round(float(ant), 2) != round(float(novo_val), 2):
                diffs.append({'campo': f'{emp} — {campo}', 'antes': _fmt_val(ant), 'depois': _fmt_val(novo_val)})
    for nome, novo_val in novo.get('BASE_FIXOS', {}).items():
        ant = antigo.get('BASE_FIXOS', {}).get(nome)
        if ant is not None and ant != novo_val:
            diffs.append({'campo': nome, 'antes': _fmt_val(ant), 'depois': _fmt_val(novo_val)})
    return diffs

def salvar_config(config: dict):
    alteracoes = _detectar_alteracoes(carregar_config(), config)
    if alteracoes:
        hist = carregar_historico()
        hist.insert(0, {'data': datetime.now().strftime('%d/%m/%Y %H:%M'), 'alteracoes': alteracoes})
        tmp_hist = HISTORY_PATH + '.tmp'
        with open(tmp_hist, 'w', encoding='utf-8') as f:
            json.dump(hist[:50], f, ensure_ascii=False, indent=2)
        os.replace(tmp_hist, HISTORY_PATH)
    tmp_cfg = CONFIG_PATH + '.tmp'
    with open(tmp_cfg, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    os.replace(tmp_cfg, CONFIG_PATH)
    print(f"[CONFIG] config.json salvo em '{CONFIG_PATH}'. {len(alteracoes)} alteração(ões) detectada(s).")
    return alteracoes

def manual_render_template(template_name, status_code=200, **kwargs):
    template_path = os.path.join(app.root_path, 'templates', template_name)
    try:
        with open(template_path, 'r', encoding='utf-8') as f:
            html_content = f.read()

        for key, value in kwargs.items():
            placeholder = f"__{key.upper()}__"
            # Trata json strings e outros tipos
            if isinstance(value, str) and value.startswith('{') and value.endswith('}'):
                 # Substitui placeholder entre aspas por valor json sem aspas
                 html_content = html_content.replace(f'"{placeholder}"', value)
            else:
                 # Substituição normal para outros tipos
                 html_content = html_content.replace(placeholder, str(value))

        response = make_response(html_content)
        response.headers['Content-Type'] = 'text/html'
        return response, status_code
    except Exception as e:
        print(f"ERRO CRÍTICO AO RENDERIZAR MANUALMENTE '{template_name}': {e}")
        # Retorna uma página de erro mais informativa
        error_html = f"""
        <!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"><title>Erro 500</title></head>
        <body><h1>Erro 500: Falha Crítica ao Carregar Template</h1>
        <p>Ocorreu um erro interno ao tentar carregar ou processar o template <strong>{template_name}</strong>.</p>
        <p><strong>Detalhes do Erro:</strong> {e}</p>
        <p>Verifique se o arquivo existe no caminho esperado ({template_path}) e se o conteúdo é válido.</p>
        </body></html>
        """
        return make_response(error_html, 500)


def normalizar_texto(s: str) -> str:
    s = s.translate(DASHES).replace("\u00A0", " ") # Substitui hífens e nbsp
    s = "".join(ch for ch in s if ch not in "\u200B\u200C\u200D\uFEFF") # Remove ZWSP e similares
    s = unicodedata.normalize("NFKC", s) # Normaliza caracteres unicode
    return s

def extrair_texto_pdf(stream_pdf) -> str:
    texto_completo = ""
    try:
        # Abre o PDF a partir do stream de bytes
        with fitz.open(stream=stream_pdf, filetype="pdf") as doc:
             # Itera sobre cada página e extrai o texto, mantendo a ordem
             for page_num in range(len(doc)):
                 page = doc.load_page(page_num)
                 # get_text("text", sort=True) tenta ordenar o texto como lido visualmente
                 texto_pagina = page.get_text("text", sort=True)
                 texto_completo += texto_pagina + "\n" # Adiciona nova linha entre páginas
        return normalizar_texto(texto_completo)
    except Exception as e:
        print(f"Erro detalhado ao ler o stream do PDF: {type(e).__name__} - {e}")
        traceback.print_exc() # Imprime o stack trace completo no log
        return "" # Retorna string vazia em caso de erro

# =======================================================
# === FUNÇÃO DE CONVERSÃO DE VALOR UNIFICADA E CORRIGIDA ===
# =======================================================
def normalizar_valor(valor):
    """Converte string para float, lidando com formatos , e . como decimal."""
    if valor is None:
        return 0.0 # Retorna 0.0 para None para simplificar somas
    if isinstance(valor, (int, float)):
        return round(float(valor), 2)
    
    s_norm = str(valor).strip().replace("R$", "").replace(" ", "").replace("\xa0", "")
    
    has_comma = "," in s_norm
    has_dot = "." in s_norm

    if has_comma and has_dot:
        # Formato 1.234,56 (vírgula é decimal)
        if s_norm.rfind(',') > s_norm.rfind('.'):
             s_norm = s_norm.replace(".", "").replace(",", ".")
        # Formato 1,234.56 (ponto é decimal)
        else:
             s_norm = s_norm.replace(",", "")
    elif has_comma:
        # Formato 1234,56 (vírgula é decimal)
        s_norm = s_norm.replace(",", ".")
    elif has_dot:
        # Formato 1234.56 (ponto é decimal)
        # OU formato 8.054.23 (pontos são milhares E decimal) - PDF INCONSISTENTE
        if s_norm.count('.') > 1:
            # Remove todos os pontos, exceto o último
            parts = s_norm.split('.')
            s_norm = "".join(parts[:-1]) + "." + parts[-1]
    
    try:
        return round(float(s_norm), 2)
    except (ValueError, TypeError):
        print(f"[AVISO] Falha ao normalizar valor: '{valor}' -> '{s_norm}'")
        return 0.0 # Retorna 0.0 em caso de falha de conversão
# =======================================================
# === FIM DA FUNÇÃO UNIFICADA ===
# =======================================================

def fixos_do_emp(emp: str, modo_separacao: str):
    """Retorna o dicionário de parcelas fixas esperadas com base no empreendimento e modo."""
    cfg = carregar_config()
    emp_map   = cfg.get('EMP_MAP', CONFIG_PADRAO['EMP_MAP'])
    base_fixos = {k: ([float(x) for x in v] if isinstance(v, list) else [float(v)]) for k, v in cfg.get('BASE_FIXOS', CONFIG_PADRAO['BASE_FIXOS']).items()}

    if modo_separacao == 'boleto':
        if emp not in emp_map:
            return base_fixos
        f = dict(base_fixos)
        if emp_map.get(emp):
            if "Melhoramentos" in emp_map[emp]:
                f["Melhoramentos"] = [float(emp_map[emp]["Melhoramentos"])]
            if "Fundo de Transporte" in emp_map[emp]:
                f["Fundo de Transporte"] = [float(emp_map[emp]["Fundo de Transporte"])]
        return f
    elif modo_separacao == 'debito_credito':
        return base_fixos
    elif modo_separacao == 'ccb_realiza':
        return BASE_FIXOS_CCB
    else:
        print(f"[AVISO] Modo de separação desconhecido '{modo_separacao}' em fixos_do_emp.")
        return {}

def detectar_emp_por_nome_arquivo(path: str):
    """Tenta detectar o código do empreendimento pelo sufixo no nome do arquivo."""
    if not path: return None
    nome = os.path.splitext(os.path.basename(path))[0].upper()
    emp_map = carregar_config().get('EMP_MAP', CONFIG_PADRAO['EMP_MAP'])
    for k in emp_map.keys():
        if nome.endswith("_" + k) or nome.endswith(k):
            return k
    # Caso especial para SBRR (se contiver no nome, mas não como sufixo exato)
    if "SBRR" in nome:
        return "SBRR" # Pode precisar de ajuste se houver outros com SBRR
    return None

def detectar_emp_por_lote(lote: str):
    """Detecta o empreendimento com base no prefixo do código do lote."""
    if not lote or "." not in lote:
        return "NAO_CLASSIFICADO"
    prefixo = lote.split('.')[0]
    # Retorna o código do mapa ou "NAO_CLASSIFICADO" se não encontrar
    return CODIGO_EMP_MAP.get(prefixo, "NAO_CLASSIFICADO")

def limpar_rotulo(lbl: str) -> str:
    """Remove prefixos e sufixos comuns dos rótulos das parcelas."""
    if not isinstance(lbl, str): return "" # Garante que é string
    lbl = re.sub(r"^TAMA\s*[-–—]\s*", "", lbl, flags=re.IGNORECASE).strip() # Remove prefixo TAMA
    lbl = re.sub(r"\s+-\s+\d+/\d+$", "", lbl).strip() # Remove sufixo de parcela N/M
    lbl = re.sub(r'\s{2,}', ' ', lbl).strip() # Remove espaços múltiplos
    return lbl

def fatiar_blocos(texto: str):
    """Divide o texto do PDF em blocos, cada um começando com um código de lote."""
    # Adiciona uma quebra de linha antes de cada padrão de lote para facilitar a divisão
    texto_processado = PADRAO_LOTE.sub(r"\n\1", texto)
    # Encontra todas as ocorrências do padrão de lote
    matches = list(PADRAO_LOTE.finditer(texto_processado))
    blocos = []
    # Itera sobre as correspondências para extrair o texto entre elas
    for i, match in enumerate(matches):
        lote_atual = match.group(1)
        inicio_bloco = match.start()
        # Fim do bloco é o início do próximo lote, ou o final do texto se for o último
        fim_bloco = matches[i+1].start() if i+1 < len(matches) else len(texto_processado)
        # Extrai o texto do bloco
        texto_bloco = texto_processado[inicio_bloco:fim_bloco].strip()
        if texto_bloco: # Adiciona apenas se o bloco não estiver vazio
             blocos.append((lote_atual, texto_bloco))
    if not blocos:
         print("[AVISO] Nenhum bloco de lote encontrado no PDF.")
    return blocos

def tentar_nome_cliente(bloco: str) -> str:
    """Tenta extrair o nome do cliente das primeiras linhas do bloco."""
    linhas = bloco.split('\n')
    if not linhas: return "Nome não localizado"

    # Considera as primeiras 5-6 linhas como candidatas
    linhas_para_buscar = linhas[:6]
    nome_candidato = "Nome não localizado"

    for linha in linhas_para_buscar:
        # Remove o código do lote da linha e espaços extras
        linha_sem_lote = PADRAO_LOTE.sub('', linha).strip()
        if not linha_sem_lote: continue # Pula linhas vazias após remover lote

        # Heurísticas mais refinadas para identificar um nome:
        is_valid_name = (
            len(linha_sem_lote) > 5 and # Pelo menos 6 caracteres
            ' ' in linha_sem_lote and # Deve conter espaço (nome composto)
            sum(c.isalpha() for c in linha_sem_lote.replace(" ", "")) / len(linha_sem_lote.replace(" ", "")) > 0.7 and # Maioria letras
            not any(h.upper() in linha_sem_lote.upper() for h in HEADERS if h) and # Não contém cabeçalhos
            not re.search(r'\d{2}/\d{2}/\d{4}', linha_sem_lote) and # Não é data
            not re.match(r'^[\d.,\s]+$', linha_sem_lote) and # Não é apenas número
            not linha_sem_lote.upper().startswith(("TOTAL", "BANCO", "03-", "LIMITE P/", "PÁGINA")) # Não começa com termos comuns
        )

        if is_valid_name:
            # Assume que a primeira linha válida encontrada é o nome
            nome_candidato = linha_sem_lote
            break # Para após encontrar o primeiro candidato válido

    return nome_candidato.strip()

def extrair_parcelas(bloco: str):
    """Extrai os nomes e valores das parcelas dentro de um bloco de texto."""
    itens = OrderedDict()
    # Tenta focar na seção "Lançamentos", se existir
    pos_lancamentos = bloco.find("Lançamentos")
    bloco_de_trabalho = bloco[pos_lancamentos:] if pos_lancamentos != -1 else bloco

    # Limpeza adicional: remove linhas de totais que podem confundir
    bloco_limpo_linhas = []
    linhas_originais = bloco_de_trabalho.splitlines()
    ignorar_proxima_linha_se_numero = False # Flag para o padrão Label \n Valor

    for i, linha in enumerate(linhas_originais):
        # Remove linhas de resumo que aparecem muito à direita
        match_total_direita = re.search(r'\s{4,}(DÉBITOS DO MÊS ANTERIOR|ENCARGOS POR ATRASO|PAGAMENTO EFETUADO)\s+[\d.,]+$', linha)
        linha_processada = linha[:match_total_direita.start()] if match_total_direita else linha
        linha_processada = linha_processada.strip()

        # Ignora linhas que são cabeçalhos conhecidos ou vazias
        if not linha_processada or any(h.strip().upper() == linha_processada.upper() for h in ["Lançamentos", "Débitos do Mês"]):
            continue

        # Se a flag estiver ativa, ignora esta linha (já foi usada como valor)
        if ignorar_proxima_linha_se_numero:
             ignorar_proxima_linha_se_numero = False
             continue

        # Tenta aplicar o padrão [Label] [Valor] na mesma linha
        match_mesma_linha = PADRAO_PARCELA_MESMA_LINHA.match(linha_processada)
        if match_mesma_linha:
            lbl = limpar_rotulo(match_mesma_linha.group(1))
            val = normalizar_valor(match_mesma_linha.group(2)) # <-- USA A FUNÇÃO CORRIGIDA
            if lbl and lbl not in itens and val is not None:
                itens[lbl] = val
                continue # Pula para a próxima linha

        # Se não casou acima, verifica se é um Label cuja próxima linha é um Valor
        is_potential_label = (
            any(c.isalpha() for c in linha_processada) and # Contém letras
            limpar_rotulo(linha_processada) not in itens # Label ainda não capturado
        )

        if is_potential_label:
            # Verifica a próxima linha NÃO VAZIA
            j = i + 1
            while j < len(linhas_originais) and not linhas_originais[j].strip():
                j += 1
            if j < len(linhas_originais):
                 linha_seguinte_limpa = linhas_originais[j].strip()
                 match_num_puro = PADRAO_NUMERO_PURO.match(linha_seguinte_limpa)
                 # Se a linha seguinte for puramente numérica
                 if match_num_puro:
                      lbl = limpar_rotulo(linha_processada)
                      val = normalizar_valor(match_num_puro.group(1)) # <-- USA A FUNÇÃO CORRIGIDA
                      if lbl and lbl not in itens and val is not None:
                           itens[lbl] = val
                           ignorar_proxima_linha_se_numero = True # Marca a linha j para ser ignorada na próxima iteração
                           continue # Pula para a próxima linha i
    return itens

def processar_pdf_validacao(texto_pdf: str, modo_separacao: str, emp_fixo_boleto: str = None):
    """Processa o texto do PDF para validação."""
    blocos = fatiar_blocos(texto_pdf)
    if not blocos: return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    linhas_todas, linhas_cov, linhas_div = [], [], []
    for lote, bloco in blocos:
        if modo_separacao == 'boleto':
            emp_atual = detectar_emp_por_lote(lote) if emp_fixo_boleto == "SBRR" else emp_fixo_boleto
        else:
            emp_atual = detectar_emp_por_lote(lote)

        cliente = tentar_nome_cliente(bloco)
        itens = extrair_parcelas(bloco)
        VALORES_CORRETOS = fixos_do_emp(emp_atual, modo_separacao) # Passa o modo

        for rot, val in itens.items():
            # 'val' já é um float corrigido pela função normalizar_valor
            linhas_todas.append({"Empreendimento": emp_atual, "Lote": lote, "Cliente": cliente, "Parcela": rot, "Valor": val})

        cov = {"Empreendimento": emp_atual, "Lote": lote, "Cliente": cliente}
        for k in VALORES_CORRETOS.keys(): cov[k] = None # Inicializa colunas
        for rot, val in itens.items():
            if rot in VALORES_CORRETOS: cov[rot] = val # Preenche valores encontrados

        vistos = [k for k in VALORES_CORRETOS if cov[k] is not None]
        cov["QtdParc_Alvo"] = len(vistos)
        cov["Parc_Alvo"] = ", ".join(vistos)
        linhas_cov.append(cov)

        # Validação de valor (apenas se houver valores permitidos definidos)
        if modo_separacao != 'ccb_realiza': # Não valida valores para CCB (lista vazia)
            for rot in vistos:
                val = cov[rot]
                if val is None: continue
                permitidos = VALORES_CORRETOS.get(rot, [])
                if permitidos and all(abs(val - v) > 1e-6 for v in permitidos):
                    linhas_div.append({
                        "Empreendimento": emp_atual, "Lote": lote, "Cliente": cliente,
                        "Parcela": rot, "Valor no Documento": float(val), # val já é float
                        "Valor Correto": " ou ".join(f"{v:.2f}" for v in permitidos)
                    })

    df_todas = pd.DataFrame(linhas_todas)
    df_cov = pd.DataFrame(linhas_cov)
    df_div = pd.DataFrame(linhas_div)

    return df_todas, df_cov, df_div

def processar_comparativo(texto_anterior, texto_atual, modo_separacao, emp_fixo_boleto):
    """Compara os dados extraídos de dois PDFs."""
    df_todas_ant_raw, _, _ = processar_pdf_validacao(texto_anterior, modo_separacao, emp_fixo_boleto)
    df_todas_atu_raw, _, _ = processar_pdf_validacao(texto_atual, modo_separacao, emp_fixo_boleto)

    # Extrai totais (agora com valores corretos de normalizar_valor)
    df_totais_ant = df_todas_ant_raw[df_todas_ant_raw['Parcela'].astype(str).str.strip().str.upper() == 'TOTAL A PAGAR'].copy()
    df_totais_ant = df_totais_ant[['Empreendimento', 'Lote', 'Cliente', 'Valor']].rename(columns={'Valor': 'Total Anterior'})

    df_totais_atu = df_todas_atu_raw[df_todas_atu_raw['Parcela'].astype(str).str.strip().str.upper() == 'TOTAL A PAGAR'].copy()
    df_totais_atu = df_totais_atu[['Empreendimento', 'Lote', 'Cliente', 'Valor']].rename(columns={'Valor': 'Total Atual'})

    # Remove parcelas indesejadas para comparação item a item
    parcelas_para_remover = ['TOTAL A PAGAR', 'DESCONTO', 'DÉBITOS DO MÊS', 'DÉBITOS DO MÊS ANTERIOR', 'ENCARGOS POR ATRASO', 'PAGAMENTO EFETUADO']
    df_todas_ant = df_todas_ant_raw[~df_todas_ant_raw['Parcela'].astype(str).str.strip().str.upper().isin(parcelas_para_remover)].copy()
    df_todas_atu = df_todas_atu_raw[~df_todas_atu_raw['Parcela'].astype(str).str.strip().str.upper().isin(parcelas_para_remover)].copy()

    df_todas_ant = df_todas_ant[~df_todas_ant['Parcela'].astype(str).str.strip().str.upper().str.startswith('TOTAL BANCO')].copy()
    df_todas_atu = df_todas_atu[~df_todas_atu['Parcela'].astype(str).str.strip().str.upper().str.startswith('TOTAL BANCO')].copy()

    df_todas_ant.rename(columns={'Valor': 'Valor Anterior'}, inplace=True)
    df_todas_atu.rename(columns={'Valor': 'Valor Atual'}, inplace=True)

    # Merge para comparação
    df_comp = pd.merge(df_todas_ant, df_todas_atu, on=['Empreendimento', 'Lote', 'Cliente', 'Parcela'], how='outer')

    # Identifica lotes adicionados/removidos
    lotes_ant = df_todas_ant_raw[['Empreendimento', 'Lote', 'Cliente']].drop_duplicates()
    lotes_atu = df_todas_atu_raw[['Empreendimento', 'Lote', 'Cliente']].drop_duplicates()
    lotes_merged = pd.merge(lotes_ant, lotes_atu, on=['Empreendimento', 'Lote', 'Cliente'], how='outer', indicator=True)

    df_adicionados_base = lotes_merged[lotes_merged['_merge'] == 'right_only'][['Empreendimento', 'Lote', 'Cliente']]
    df_removidos_base = lotes_merged[lotes_merged['_merge'] == 'left_only'][['Empreendimento', 'Lote', 'Cliente']]

    # Adiciona valor total aos lotes adicionados/removidos
    df_adicionados = pd.merge(df_adicionados_base, df_totais_atu, on=['Empreendimento', 'Lote', 'Cliente'], how='left')
    df_removidos = pd.merge(df_removidos_base, df_totais_ant, on=['Empreendimento', 'Lote', 'Cliente'], how='left')

    # Identifica divergências de valor, parcelas novas e removidas
    df_divergencias = df_comp[
        (pd.notna(df_comp['Valor Anterior'])) &
        (pd.notna(df_comp['Valor Atual'])) &
        (abs(df_comp['Valor Anterior'] - df_comp['Valor Atual']) > 0.025) # Tolerância de ~2 centavos
    ].copy()
    if not df_divergencias.empty:
         df_divergencias['Diferença'] = df_divergencias['Valor Atual'] - df_divergencias['Valor Anterior']

    df_parcelas_novas = df_comp[df_comp['Valor Anterior'].isna() & pd.notna(df_comp['Valor Atual'])][['Empreendimento', 'Lote', 'Cliente', 'Parcela', 'Valor Atual']].copy()
    df_parcelas_removidas = df_comp[df_comp['Valor Atual'].isna() & pd.notna(df_comp['Valor Anterior'])][['Empreendimento', 'Lote', 'Cliente', 'Parcela', 'Valor Anterior']].copy()

    # Calcula totais para o resumo
    total_adicionados_valor = df_adicionados['Total Atual'].sum() if 'Total Atual' in df_adicionados.columns else 0
    total_removidos_valor = df_removidos['Total Anterior'].sum() if 'Total Anterior' in df_removidos.columns else 0
    total_divergencias_valor = df_divergencias['Diferença'].sum() if 'Diferença' in df_divergencias.columns else 0
    total_mes_anterior_valor = df_totais_ant['Total Anterior'].sum() if 'Total Anterior' in df_totais_ant.columns else 0
    total_mes_atual_valor = df_totais_atu['Total Atual'].sum() if 'Total Atual' in df_totais_atu.columns else 0

    # Cria DataFrame de resumo
    resumo_financeiro_data = {
        ' ': ['Lotes Mês Anterior', 'Lotes Mês Atual', 'Lotes Adicionados', 'Lotes Removidos', 'Parcelas com Valor Alterado'],
        'LOTES': [len(lotes_ant), len(lotes_atu), len(df_adicionados), len(df_removidos), df_divergencias['Lote'].nunique() if not df_divergencias.empty else 0],
        'TOTAIS': [total_mes_anterior_valor, total_mes_atual_valor, total_adicionados_valor, total_removidos_valor, total_divergencias_valor]
    }
    df_resumo_completo = pd.DataFrame(resumo_financeiro_data)

    # Retorna todos os DataFrames gerados
    return df_resumo_completo, df_adicionados, df_removidos, df_divergencias, df_parcelas_novas, df_parcelas_removidas


def formatar_excel(output_stream, dfs: dict):
    """Formata planilhas de Validação e Comparação (não Repasse)."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    import pandas as pd

    with pd.ExcelWriter(output_stream, engine='openpyxl') as writer:
        for sheet_name, df in dfs.items():
            if df is None:
                continue
            if isinstance(df, pd.DataFrame):
                 df.to_excel(writer, index=False, sheet_name=sheet_name)
            else:
                 print(f"[AVISO] Tentando salvar algo que não é DataFrame na planilha '{sheet_name}': {type(df)}")
                 pd.DataFrame([{"Erro": f"Dados inválidos para {sheet_name}"}]).to_excel(writer, index=False, sheet_name=sheet_name)

        number_style = NamedStyle(name='br_number_style', number_format='#,##0.00')
        integer_style = NamedStyle(name='br_integer_style', number_format='0')

        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            worksheet.sheet_view.showGridLines = False

            if worksheet.max_row > 0 and worksheet.max_column > 0:
                for col_idx, column_cells in enumerate(worksheet.columns, 1):
                    max_length = 0
                    column = get_column_letter(col_idx)
                    is_first_row = True
                    for cell in column_cells:
                        if cell.value:
                             try:
                                 if isinstance(cell.value, (int, float)) and cell.number_format != 'General' and not is_first_row:
                                     if '##0.00' in cell.number_format:
                                         formatted_value = f"{cell.value:,.2f}"
                                     elif '0' == cell.number_format:
                                         formatted_value = f"{cell.value:d}"
                                     else:
                                         formatted_value = str(cell.value)
                                     max_length = max(max_length, len(formatted_value))
                                 else:
                                      max_length = max(max_length, len(str(cell.value)))
                             except:
                                  max_length = max(max_length, len(str(cell.value)))


                        if not is_first_row:
                             if isinstance(cell.value, float):
                                 cell.style = number_style
                             elif isinstance(cell.value, int):
                                 if sheet_name == 'Resumo' and column == 'B':
                                     cell.style = integer_style
                                 elif column != 'B':
                                     cell.style = number_style
                        is_first_row = False

                    adjusted_width = (max_length + 2) * 1.15
                    worksheet.column_dimensions[column].width = min(max(adjusted_width, 10), 60)

                worksheet.auto_filter.ref = worksheet.dimensions
                print(f"[LOG] Autofilter aplicado à planilha '{sheet_name}'. Ref: {worksheet.dimensions}")
    return output_stream


# ==== ROTAS FLASK ====

@app.route('/')
def index():
    return manual_render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'pdf_file' not in request.files or request.files['pdf_file'].filename == '':
        return manual_render_template('error.html', status_code=400,
            error_title="Nenhum arquivo enviado",
            error_message="Você precisa selecionar um arquivo PDF para fazer a análise.")

    file = request.files['pdf_file']
    modo_separacao = request.form.get('modo_separacao', 'boleto')

    try:
        emp_fixo = None
        if modo_separacao == 'boleto':
            emp_fixo = detectar_emp_por_nome_arquivo(file.filename)
            if not emp_fixo:
                error_msg = ("Para o modo 'Boleto', o nome do arquivo precisa terminar com um código de empreendimento válido (ex: 'Extrato_RSCI.pdf'). "
                             "Verifique o nome do arquivo ou selecione outro modo de análise.")
                return manual_render_template('error.html', status_code=400,
                    error_title="Empreendimento não identificado (Modo Boleto)", error_message=error_msg)

        elif modo_separacao in ['debito_credito', 'ccb_realiza']:
             if detectar_emp_por_nome_arquivo(file.filename) and modo_separacao == 'debito_credito':
                  error_msg = ("Este arquivo parece ser do tipo 'Boleto' (termina com código de empreendimento), mas o modo 'Débito/Crédito' foi selecionado. "
                               "Por favor, use o modo 'Boleto' ou renomeie o arquivo se ele não for específico de um empreendimento.")
                  return manual_render_template('error.html', status_code=400,
                                                error_title="Modo de Análise Incorreto?", error_message=error_msg)

        print(f"Iniciando validação para o arquivo '{file.filename}' no modo '{modo_separacao}'...")
        pdf_stream = file.read()
        texto_pdf = extrair_texto_pdf(pdf_stream)
        if not texto_pdf:
            print(f"Falha ao extrair texto do PDF: {file.filename}")
            return manual_render_template('error.html', status_code=500,
                error_title="Erro ao ler o PDF",
                error_message="Não foi possível extrair o texto do arquivo enviado. Ele pode estar corrompido, ser uma imagem ou estar vazio.")

        print("Texto extraído, processando validação...")
        df_todas_raw, df_cov, df_div = processar_pdf_validacao(texto_pdf, modo_separacao, emp_fixo)
        print(f"Validação concluída. {len(df_cov)} lotes/registros encontrados, {len(df_div)} divergências.")

        df_todas_filtrado = df_todas_raw.copy()
        if not df_todas_filtrado.empty:
            parcelas_para_remover = ['TOTAL A PAGAR', 'DESCONTO', 'DÉBITOS DO MÊS ANTERIOR', 'ENCARGOS POR ATRASO', 'PAGAMENTO EFETUADO', 'DÉBITOS DO MÊS']
            df_todas_filtrado = df_todas_filtrado[~df_todas_filtrado['Parcela'].astype(str).str.strip().str.upper().isin(parcelas_para_remover)]
            df_todas_filtrado = df_todas_filtrado[~df_todas_filtrado['Parcela'].astype(str).str.strip().str.upper().str.startswith('TOTAL BANCO')]
        print("Parcelas indesejadas filtradas da aba 'Todas_Parcelas_Extraidas'.")

        output = io.BytesIO()
        dfs_to_excel = {"Divergencias": df_div, "Cobertura_Analise": df_cov, "Todas_Parcelas_Extraidas": df_todas_filtrado}
        print("Gerando arquivo Excel...")
        formatar_excel(output, dfs_to_excel) # Chama a função formatar_excel com autofiltro
        output.seek(0)
        print("Arquivo Excel gerado em memória.")

        base_name = os.path.splitext(file.filename)[0]
        report_filename = f"relatorio_{modo_separacao}_{base_name}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(app.config['UPLOAD_FOLDER'], report_filename)

        try:
            with open(report_path, 'wb') as f: f.write(output.getvalue())
            print(f"Relatório salvo em: {report_path}")
        except Exception as e_save:
            print(f"Erro ao salvar o arquivo Excel em {report_path}: {e_save}")

        nao_classificados = 0
        if not df_cov.empty and 'Empreendimento' in df_cov.columns:
            nao_classificados = df_cov[df_cov['Empreendimento'] == 'NAO_CLASSIFICADO'].shape[0]
            if nao_classificados > 0: print(f"[AVISO] {nao_classificados} registros não classificados.")

        print("Renderizando página de resultados...")
        return manual_render_template('results.html',
            divergencias_json=df_div.to_json(orient='split', index=False, date_format='iso') if not df_div.empty else 'null',
            total_lotes=len(df_cov),
            total_divergencias=len(df_div),
            nao_classificados=nao_classificados,
            download_url=url_for('download_file', filename=report_filename),
            modo_usado=modo_separacao.replace('_', '/').upper()
        )

    except Exception as e:
        print(f"📕 [ERRO FATAL] Erro inesperado na rota /upload: {e}")
        traceback.print_exc()
        return manual_render_template('error.html', status_code=500,
            error_title="Erro inesperado no processamento",
            error_message=f"Ocorreu um erro grave durante a análise do arquivo '{file.filename}'. Detalhes: {e}")

@app.route('/compare', methods=['POST'])
def compare_files():
    if 'pdf_mes_anterior' not in request.files or 'pdf_mes_atual' not in request.files:
        return manual_render_template('error.html', status_code=400,
            error_title="Arquivos faltando",
            error_message="Ambos os arquivos PDF (mês anterior e atual) são necessários para a comparação.")

    file_ant = request.files['pdf_mes_anterior']
    file_atu = request.files['pdf_mes_atual']
    modo_separacao = request.form.get('modo_separacao_comp', 'boleto')

    if file_ant.filename == '' or file_atu.filename == '':
        return manual_render_template('error.html', status_code=400,
            error_title="Arquivos faltando",
            error_message="Selecione os dois arquivos PDF para comparar.")

    if not file_ant.filename.lower().endswith('.pdf') or not file_atu.filename.lower().endswith('.pdf'):
         return manual_render_template('error.html', status_code=400,
            error_title="Tipo de Arquivo Inválido",
            error_message="Por favor, envie apenas arquivos no formato PDF para comparação.")


    try:
        emp_fixo_boleto = None
        if modo_separacao == 'boleto':
            emp_ant = detectar_emp_por_nome_arquivo(file_ant.filename)
            emp_atu = detectar_emp_por_nome_arquivo(file_atu.filename)
            if not emp_ant or not emp_atu:
                return manual_render_template('error.html', status_code=400,
                    error_title="Empreendimento não identificado (Modo Boleto)",
                    error_message="Para o modo 'Boleto', o nome de ambos os arquivos PDF precisa terminar com um código de empreendimento válido.")
            if emp_ant != emp_atu:
                return manual_render_template('error.html', status_code=400,
                    error_title="Empreendimentos diferentes (Modo Boleto)",
                    error_message=f"Os arquivos devem ser do mesmo empreendimento para comparação no modo Boleto (Detectado: '{emp_ant}' e '{emp_atu}').")
            emp_fixo_boleto = emp_ant

        elif modo_separacao in ['debito_credito', 'ccb_realiza']:
             if detectar_emp_por_nome_arquivo(file_ant.filename) or detectar_emp_por_nome_arquivo(file_atu.filename):
                  error_msg = (f"Um dos arquivos parece ser do tipo 'Boleto' (termina com código), mas o modo '{modo_separacao.replace('_','/').upper()}' foi selecionado. "
                               "Use o modo 'Boleto' para esses arquivos ou renomeie-os se a detecção estiver incorreta.")
                  return manual_render_template('error.html', status_code=400,
                                                error_title="Modo de Análise Incorreto?", error_message=error_msg)

        print(f"Iniciando comparação modo '{modo_separacao}' entre '{file_ant.filename}' e '{file_atu.filename}'...")
        texto_ant = extrair_texto_pdf(file_ant.read())
        texto_atu = extrair_texto_pdf(file_atu.read())

        if not texto_ant or not texto_atu:
            err_msg = "Não foi possível extrair texto de um ou ambos os PDFs. "
            if not texto_ant and not texto_atu: err_msg += "Ambos os arquivos falharam."
            elif not texto_ant: err_msg += f"Falha ao ler '{file_ant.filename}'."
            else: err_msg += f"Falha ao ler '{file_atu.filename}'."
            err_msg += " Verifique se não estão corrompidos ou se são imagens."
            print(f"[ERRO] {err_msg}")
            return manual_render_template('error.html', status_code=500,
                error_title="Erro ao ler PDF na Comparação", error_message=err_msg)

        print("Textos extraídos. Processando comparação...")
        df_resumo_completo, df_adicionados, df_removidos, df_divergencias, df_parcelas_novas, df_parcelas_removidas = processar_comparativo(
            texto_ant, texto_atu, modo_separacao, emp_fixo_boleto
        )
        print(f"Comparação concluída. Resumo: {len(df_adicionados)} adicionados, {len(df_removidos)} removidos, {len(df_divergencias)} divergências.")


        output = io.BytesIO()
        dfs_to_excel = {
            "Resumo": df_resumo_completo,
            "Lotes Adicionados": df_adicionados,
            "Lotes Removidos": df_removidos,
            "Divergências de Valor": df_divergencias,
            "Parcelas Novas por Lote": df_parcelas_novas,
            "Parcelas Removidas por Lote": df_parcelas_removidas,
        }
        print("Gerando arquivo Excel do comparativo...")
        formatar_excel(output, dfs_to_excel)
        output.seek(0)
        print("Arquivo Excel gerado em memória.")

        report_filename = f"comparativo_{modo_separacao}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(app.config['UPLOAD_FOLDER'], report_filename)
        try:
            with open(report_path, 'wb') as f:
                f.write(output.getvalue())
            print(f"Relatório comparativo salvo em: {report_path}")
        except Exception as e_save:
             print(f"Erro ao salvar o arquivo Excel comparativo em {report_path}: {e_save}")


        resumo_dict_lotes = {}
        resumo_dict_totais = {}
        if not df_resumo_completo.empty:
             resumo_dict_lotes = pd.Series(df_resumo_completo.set_index(' ')['LOTES']).to_dict()
             resumo_dict_totais = pd.Series(df_resumo_completo.set_index(' ')['TOTAIS']).map('{:,.2f}'.format).to_dict()


        print("Renderizando página de resultados da comparação...")
        return manual_render_template('compare_results.html',
             resumo_lotes_mes_anterior=resumo_dict_lotes.get('Lotes Mês Anterior', 0),
             resumo_lotes_mes_atual=resumo_dict_lotes.get('Lotes Mês Atual', 0),
             resumo_lotes_adicionados=resumo_dict_lotes.get('Lotes Adicionados', 0),
             resumo_lotes_removidos=resumo_dict_lotes.get('Lotes Removidos', 0),
             resumo_parcelas_com_valor_alterado=resumo_dict_lotes.get('Parcelas com Valor Alterado', 0),

             total_mes_anterior_str=resumo_dict_totais.get('Lotes Mês Anterior', '0.00'),
             total_mes_atual_str=resumo_dict_totais.get('Lotes Mês Atual', '0.00'),
             total_adicionados_str=resumo_dict_totais.get('Lotes Adicionados', '0.00'),
             total_removidos_str=resumo_dict_totais.get('Lotes Removidos', '0.00'),
             total_diferencas_str=resumo_dict_totais.get('Parcelas com Valor Alterado', '0.00'),

            divergencias_json=df_divergencias.to_json(orient='split', index=False, date_format='iso') if not df_divergencias.empty else 'null',
            adicionados_json=df_adicionados.to_json(orient='split', index=False, date_format='iso') if not df_adicionados.empty else 'null',
            removidos_json=df_removidos.to_json(orient='split', index=False, date_format='iso') if not df_removidos.empty else 'null',

            download_url=url_for('download_file', filename=report_filename),
            modo_usado=modo_separacao.replace('_', '/').upper()
        )


    except Exception as e:
        print(f"📕 [ERRO FATAL] Erro inesperado na rota /compare: {e}")
        traceback.print_exc()
        error_details = f"{type(e).__name__}: {e}"
        return manual_render_template('error.html', status_code=500,
            error_title="Erro inesperado na comparação",
            error_message=f"Ocorreu um erro grave durante a comparação dos arquivos. Detalhes: {error_details}")


@app.route('/configuracoes/login', methods=['GET', 'POST'])
def configuracoes_login():
    erro_html = ''
    if request.method == 'POST':
        if request.form.get('senha') == CONFIG_SENHA:
            session['config_auth'] = True
            return redirect('/configuracoes')
        erro_html = '<div class="alert-erro">Senha incorreta. Verifique e tente novamente.</div>'
    return manual_render_template('configuracoes_login.html', erro=erro_html)

@app.route('/configuracoes')
def configuracoes():
    if not session.get('config_auth'):
        return redirect('/configuracoes/login')
    cfg = carregar_config()
    return manual_render_template(
        'configuracoes.html',
        emp_map=json.dumps(cfg.get('EMP_MAP', CONFIG_PADRAO['EMP_MAP']), ensure_ascii=False),
        base_fixos=json.dumps(cfg.get('BASE_FIXOS', CONFIG_PADRAO['BASE_FIXOS']), ensure_ascii=False),
    )

@app.route('/configuracoes/historico')
def configuracoes_historico():
    if not session.get('config_auth'):
        return jsonify([])
    return jsonify(carregar_historico())

@app.route('/configuracoes/sair')
def configuracoes_sair():
    session.pop('config_auth', None)
    return redirect('/')

@app.route('/configuracoes/salvar', methods=['POST'])
def configuracoes_salvar():
    if not session.get('config_auth'):
        return jsonify({'ok': False, 'erro': 'Não autorizado.'}), 401
    try:
        data = request.get_json(force=True)
        emp_map = {k: {"Melhoramentos": float(v["Melhoramentos"]), "Fundo de Transporte": float(v["Fundo de Transporte"])} for k, v in data.get('EMP_MAP', {}).items()}
        base_fixos = {}
        for k, v in data.get('BASE_FIXOS', {}).items():
            vals = [float(x) for x in v if x is not None and float(x) > 0] if isinstance(v, list) else ([float(v)] if v else [])
            base_fixos[k] = vals
        nova_config = {"EMP_MAP": emp_map, "BASE_FIXOS": base_fixos}
        alteracoes = salvar_config(nova_config)
        # Verificação pós-gravação
        cfg_lido = carregar_config()
        ok = cfg_lido.get('EMP_MAP') == emp_map
        print(f"[CONFIG] Verificação pós-save: EMP_MAP correto={ok}, arquivo existe={os.path.exists(CONFIG_PATH)}")
        # Commit automático no GitHub (torna a mudança permanente)
        github_ok, github_msg = False, 'GitHub não configurado.'
        if github_configurado():
            github_ok, github_msg = commitar_config_github(nova_config, alteracoes)
        return jsonify({'ok': True, 'github_ok': github_ok, 'github_msg': github_msg})
    except Exception as e:
        print(f"[CONFIG] ERRO ao salvar: {e}")
        return jsonify({'ok': False, 'erro': str(e)}), 500

@app.route('/configuracoes/verificar')
def configuracoes_verificar():
    if not session.get('config_auth'):
        return jsonify({'erro': 'Não autorizado.'}), 401
    cfg = carregar_config()
    return jsonify({
        'config_path': CONFIG_PATH,
        'arquivo_existe': os.path.exists(CONFIG_PATH),
        'tamanho_bytes': os.path.getsize(CONFIG_PATH) if os.path.exists(CONFIG_PATH) else 0,
        'config': cfg,
    })

@app.route('/download/<filename>')
def download_file(filename):
     safe_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
     normalized_safe_path = os.path.normpath(safe_path)
     normalized_upload_folder = os.path.normpath(app.config['UPLOAD_FOLDER'])

     # Adiciona 'os.sep' para garantir que não pegue pastas com nome parecido
     if not normalized_safe_path.startswith(normalized_upload_folder + os.sep) and normalized_safe_path != normalized_upload_folder :
         print(f" Tentativa de acesso a caminho inválido: {filename} (Normalizado: {normalized_safe_path} vs Base: {normalized_upload_folder})")
         return "Acesso negado.", 403

     if not os.path.exists(safe_path):
          print(f" Arquivo não encontrado para download: {filename}")
          return "Arquivo não encontrado.", 404

     print(f"Enviando arquivo para download: {filename}")
     return send_file(safe_path, as_attachment=True)


if __name__ == '__main__':
    print("Iniciando servidor Flask local...")
    port = int(os.environ.get('PORT', 8080))
    # Verifica variável de ambiente FLASK_DEBUG para modo debug
    debug_mode = os.environ.get('FLASK_DEBUG') == '1'
    # Usa host='0.0.0.0' para ser acessível na rede local ou pelo Render
    print(f"Executando em http://0.0.0.0:{port} (debug={debug_mode})")
    # threaded=True pode ajudar a evitar timeouts em requisições longas localmente
    app.run(debug=debug_mode, host='0.0.0.0', port=port, threaded=True)

